#!/usr/bin/env python3
"""
Complete OCR Pipeline - Combines Scripts 01, 03, and 04
Processes PDF/images through the entire pipeline: PDF→PNG→Gemini→Excel
"""

import os
import sys
import json
import logging
import argparse
import base64
import requests
import io
import time
import random
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import fitz  # PyMuPDF
from PIL import Image
import pandas as pd
from dotenv import load_dotenv

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class CompletePipeline:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.gemini_url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent"
        self.headers = {"Content-Type": "application/json"}
    
    def pdf_to_images(self, pdf_path: str, output_dir: str, rotation: int = 0, dpi: int = 300, max_pages: int = None) -> List[str]:
        """Convert PDF to PNG images with optional rotation"""
        logger.info(f"🔄 Converting PDF: {os.path.basename(pdf_path)}")
        
        # Create output directory
        os.makedirs(output_dir, exist_ok=True)
        
        # Open PDF
        doc = fitz.open(pdf_path)
        image_paths = []
        
        # Limit pages if max_pages is specified
        total_pages = len(doc)
        pages_to_process = min(total_pages, max_pages) if max_pages else total_pages
        
        logger.info(f"   📄 Processing {pages_to_process} of {total_pages} pages")
        
        for page_num in range(pages_to_process):
            page = doc.load_page(page_num)
            
            # Convert to image
            mat = fitz.Matrix(dpi/72, dpi/72)
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            
            # Convert to PIL Image for rotation
            img = Image.open(io.BytesIO(img_data))
            
            # Apply rotation if specified
            if rotation != 0:
                img = img.rotate(-rotation, expand=True)
            
            # Save image
            page_filename = f"page_{page_num+1:03d}.png"
            page_path = os.path.join(output_dir, page_filename)
            img.save(page_path)
            image_paths.append(page_path)
            
            logger.info(f"   ✅ Page {page_num+1} → {page_filename}")
        
        doc.close()
        logger.info(f"📄 Converted {len(image_paths)} pages")
        return image_paths
    
    def process_single_image(self, image_path: str, rotation: int = 0) -> str:
        """Process a single image file with optional rotation"""
        logger.info(f"🖼️ Processing image: {os.path.basename(image_path)}")
        
        if rotation == 0:
            return image_path
        
        # Create output path for rotated image
        output_dir = os.path.dirname(image_path)
        base_name = os.path.splitext(os.path.basename(image_path))[0]
        rotated_path = os.path.join(output_dir, f"{base_name}_rotated.png")
        
        # Apply rotation
        img = Image.open(image_path)
        img_rotated = img.rotate(-rotation, expand=True)
        img_rotated.save(rotated_path)
        
        logger.info(f"   🔄 Rotated and saved: {os.path.basename(rotated_path)}")
        return rotated_path
    
    def extract_table_with_gemini(self, image_path: str) -> dict:
        """Extract table data using Gemini API with retry logic"""
        logger.info(f"🔍 Gemini extraction: {os.path.basename(image_path)}")
        
        max_retries = 3
        base_delay = 2
        
        for attempt in range(max_retries):
            try:
                # Encode image
                with open(image_path, "rb") as f:
                    image_data = f.read()
                    base64_image = base64.b64encode(image_data).decode('utf-8')
                
                # Gemini prompt
                prompt = """
                You are an expert at analyzing table images and extracting structured data.

                Analyze this image and extract the table data in the following JSON format:

                {
                    "table_info": {
                        "total_rows": <number>,
                        "total_columns": <number>,
                        "headers": ["header1", "header2", "header3", ...]
                    },
                    "data": [
                        {
                            "row": 1,
                            "cells": ["cell1_value", "cell2_value", "cell3_value", ...]
                        },
                        {
                            "row": 2,
                            "cells": ["cell1_value", "cell2_value", "cell3_value", ...]
                        }
                    ]
                }

                IMPORTANT INSTRUCTIONS:
                1. Look carefully at the image and identify ALL table content
                2. Extract column headers from the first row
                3. Extract ALL data from each row
                4. Be very accurate with text extraction
                5. Return ONLY the JSON, no other text
                6. If you see handwritten text, transcribe it accurately
                7. Include all visible information in the table

                Return the data in the exact JSON format shown above.
                """
                
                # API request
                payload = {
                    "contents": [{
                        "parts": [
                            {"text": prompt},
                            {
                                "inline_data": {
                                    "mime_type": "image/png",
                                    "data": base64_image
                                }
                            }
                        ]
                    }],
                    "generationConfig": {
                        "temperature": 0.1,
                        "topK": 1,
                        "topP": 1,
                        "maxOutputTokens": 8192
                    }
                }
                
                response = requests.post(
                    f"{self.gemini_url}?key={self.api_key}",
                    headers=self.headers,
                    json=payload
                )
                
                if response.status_code == 200:
                    # Success - parse response
                    response_data = response.json()
                    if 'candidates' in response_data and len(response_data['candidates']) > 0:
                        content = response_data['candidates'][0]['content']
                        if 'parts' in content and len(content['parts']) > 0:
                            extracted_text = content['parts'][0]['text']
                            return self.parse_gemini_response(extracted_text)
                    
                    return {"table_info": {"total_rows": 0, "total_columns": 0, "headers": []}, "data": []}
                
                elif response.status_code in [503, 429, 500, 502, 504]:
                    # Retryable errors (rate limit, server issues)
                    if attempt < max_retries - 1:
                        delay = base_delay * (2 ** attempt) + random.uniform(0, 1)
                        logger.warning(f"   ⚠️ API error {response.status_code}, retrying in {delay:.1f}s (attempt {attempt+1}/{max_retries})")
                        time.sleep(delay)
                        continue
                    else:
                        logger.error(f"❌ Gemini API failed after {max_retries} attempts: {response.status_code}")
                        return {"table_info": {"total_rows": 0, "total_columns": 0, "headers": []}, "data": []}
                else:
                    # Non-retryable error
                    logger.error(f"❌ Gemini API failed: {response.status_code}")
                    logger.error(f"   Response: {response.text[:200]}...")
                    return {"table_info": {"total_rows": 0, "total_columns": 0, "headers": []}, "data": []}
                    
            except Exception as e:
                if attempt < max_retries - 1:
                    delay = base_delay * (2 ** attempt)
                    logger.warning(f"   ⚠️ Exception: {str(e)}, retrying in {delay}s")
                    time.sleep(delay)
                    continue
                else:
                    logger.error(f"❌ Gemini extraction failed: {e}")
                    return {"table_info": {"total_rows": 0, "total_columns": 0, "headers": []}, "data": []}
        
        return {"table_info": {"total_rows": 0, "total_columns": 0, "headers": []}, "data": []}

    def parse_gemini_response(self, response_text: str) -> dict:
        """Parse Gemini response to extract JSON table data"""
        try:
            # Look for JSON in the response
            start_idx = response_text.find('{')
            end_idx = response_text.rfind('}') + 1
            
            if start_idx != -1 and end_idx != -1:
                json_str = response_text[start_idx:end_idx]
                table_data = json.loads(json_str)
                return table_data
            else:
                return {"table_info": {"total_rows": 0, "total_columns": 0, "headers": []}, "data": []}
                
        except json.JSONDecodeError:
            return {"table_info": {"total_rows": 0, "total_columns": 0, "headers": []}, "data": []}
    
    def create_dataframe_from_gemini(self, gemini_data: dict) -> pd.DataFrame:
        """Convert Gemini data to pandas DataFrame"""
        try:
            headers = gemini_data.get('table_info', {}).get('headers', [])
            rows_data = gemini_data.get('data', [])
            
            if not headers or not rows_data:
                return pd.DataFrame()
            
            # Create DataFrame
            df_data = []
            for row in rows_data:
                cells = row.get('cells', [])
                # Ensure we have the right number of columns
                if len(cells) == len(headers):
                    df_data.append(cells)
                else:
                    # Pad with empty strings if columns don't match
                    padded_cells = cells + [''] * (len(headers) - len(cells))
                    df_data.append(padded_cells[:len(headers)])
            
            df = pd.DataFrame(df_data, columns=headers)
            
            # Clean the data
            df.columns = [col.strip() for col in df.columns]
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].astype(str).str.strip()
                    df[col] = df[col].replace('', pd.NA)
            
            df = df.dropna(how='all').reset_index(drop=True)
            return df
            
        except Exception as e:
            logger.error(f"❌ Error creating DataFrame: {e}")
            return pd.DataFrame()
    
    def save_to_excel(self, df: pd.DataFrame, output_path: str):
        """Save DataFrame to Excel with formatting"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Table Data', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Table Data']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format header row
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
                    for cell in row:
                        cell.border = thin_border
            
            logger.info(f"💾 Excel saved: {output_path}")
            
        except Exception as e:
            logger.error(f"❌ Error saving Excel: {e}")
    
    def process_files(self, input_paths: List[str], output_dir: str, rotation: int = 0, dpi: int = 300, max_pages: int = None) -> pd.DataFrame:
        """Process multiple files and combine results"""
        os.makedirs(output_dir, exist_ok=True)
        
        all_dataframes = []
        processed_images = []
        
        for input_path in input_paths:
            logger.info(f"📁 Processing: {input_path}")
            
            if input_path.lower().endswith('.pdf'):
                # PDF processing
                pdf_output_dir = os.path.join(output_dir, f"{Path(input_path).stem}_images")
                image_paths = self.pdf_to_images(input_path, pdf_output_dir, rotation, dpi, max_pages)
                processed_images.extend(image_paths)
            
            elif input_path.lower().endswith(('.png', '.jpg', '.jpeg')):
                # Image processing
                processed_image = self.process_single_image(input_path, rotation)
                processed_images.append(processed_image)
            
            else:
                logger.warning(f"⚠️ Unsupported file type: {input_path}")
                continue
        
        # Extract tables from all images
        logger.info(f"🔍 Extracting tables from {len(processed_images)} images...")
        
        for i, image_path in enumerate(processed_images, 1):
            logger.info(f"[{i}/{len(processed_images)}] Processing: {os.path.basename(image_path)}")
            
            gemini_data = self.extract_table_with_gemini(image_path)
            df = self.create_dataframe_from_gemini(gemini_data)
            
            if not df.empty:
                logger.info(f"   ✅ Extracted: {df.shape[0]} rows × {df.shape[1]} columns")
                all_dataframes.append(df)
            else:
                logger.info(f"   ⚠️ No table data found")
        
        if not all_dataframes:
            logger.warning("❌ No tables extracted from any images")
            return pd.DataFrame()
        
        # Combine all DataFrames
        logger.info(f"📊 Combining {len(all_dataframes)} tables...")
        
        if len(all_dataframes) == 1:
            combined_df = all_dataframes[0]
        else:
            # Align columns across all DataFrames
            all_columns = set()
            for df in all_dataframes:
                all_columns.update(df.columns)
            all_columns = sorted(list(all_columns))
            
            # Ensure all DataFrames have the same columns
            aligned_dfs = []
            for df in all_dataframes:
                for col in all_columns:
                    if col not in df.columns:
                        df[col] = ""
                aligned_dfs.append(df[all_columns])
            
            combined_df = pd.concat(aligned_dfs, ignore_index=True)
        
        logger.info(f"🎉 Combined result: {combined_df.shape[0]} rows × {combined_df.shape[1]} columns")
        return combined_df

def main():
    parser = argparse.ArgumentParser(description='Complete OCR Pipeline: PDF/Images → Gemini → Excel')
    parser.add_argument('inputs', nargs='+', help='Input files (PDF, PNG, JPG)')
    parser.add_argument('--api-key', help='Gemini API key (optional if GEMINI_API_KEY in .env)')
    parser.add_argument('--output', '-o', default='data/output/complete_pipeline', help='Output directory')
    parser.add_argument('--rotation', type=int, default=0, choices=[0, 90, 180, 270], help='Rotate all images (degrees)')
    parser.add_argument('--dpi', type=int, default=300, help='DPI for PDF conversion')
    parser.add_argument('--max-pages', type=int, help='Maximum number of pages to process (useful for testing)')
    
    args = parser.parse_args()
    
    # Load environment variables
    load_dotenv()
    
    # Get API key
    api_key = args.api_key or os.getenv('GEMINI_API_KEY')
    if not api_key:
        logger.error("❌ Gemini API key required (--api-key or GEMINI_API_KEY in .env)")
        sys.exit(1)
    
    # Validate input files
    for input_path in args.inputs:
        if not os.path.exists(input_path):
            logger.error(f"❌ File not found: {input_path}")
            sys.exit(1)
    
    try:
        # Initialize pipeline
        pipeline = CompletePipeline(api_key)
        
        # Process files
        logger.info(f"🚀 Starting Complete OCR Pipeline...")
        logger.info(f"   Input files: {len(args.inputs)}")
        logger.info(f"   Output directory: {args.output}")
        logger.info(f"   Rotation: {args.rotation}°")
        
        combined_df = pipeline.process_files(args.inputs, args.output, args.rotation, args.dpi, args.max_pages)
        
        if combined_df.empty:
            logger.error("❌ No data extracted")
            sys.exit(1)
        
        # Save results
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"complete_extraction_{timestamp}.xlsx"
        csv_filename = f"complete_extraction_{timestamp}.csv"
        json_filename = f"complete_extraction_{timestamp}.json"
        
        excel_path = os.path.join(args.output, excel_filename)
        csv_path = os.path.join(args.output, csv_filename)
        json_path = os.path.join(args.output, json_filename)
        
        # Save in multiple formats
        pipeline.save_to_excel(combined_df, excel_path)
        combined_df.to_csv(csv_path, index=False)
        combined_df.to_json(json_path, orient='records', indent=2)
        
        logger.info(f"💾 Results saved:")
        logger.info(f"   📊 Excel: {excel_filename}")
        logger.info(f"   📄 CSV: {csv_filename}")
        logger.info(f"   📋 JSON: {json_filename}")
        
        # Summary
        logger.info(f"\n🎉 Pipeline Complete!")
        logger.info(f"   Total rows: {combined_df.shape[0]}")
        logger.info(f"   Total columns: {combined_df.shape[1]}")
        logger.info(f"   Headers: {', '.join(combined_df.columns[:5])}{'...' if len(combined_df.columns) > 5 else ''}")
        
    except Exception as e:
        logger.error(f"❌ Pipeline failed: {e}")
        raise

if __name__ == "__main__":
    main()
