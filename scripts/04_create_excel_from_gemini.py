#!/usr/bin/env python3
"""
Script 4 (Gemini): Create Excel Table from Gemini API Data
Creates professional Excel table from Gemini API extracted data.
"""

import os
import sys
import json
import logging
import argparse
import pandas as pd
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def load_gemini_data(gemini_data_path: str):
    """
    Load Gemini API extracted table data.
    """
    try:
        logger.info(f"📂 Loading Gemini API data from: {os.path.basename(gemini_data_path)}")
        
        with open(gemini_data_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        logger.info(f"   ✅ Successfully loaded Gemini data")
        return data
        
    except Exception as e:
        logger.error(f"❌ Error loading Gemini data: {e}")
        raise

def create_dataframe_from_gemini(gemini_data: dict):
    """
    Convert Gemini API data to pandas DataFrame.
    """
    try:
        logger.info("   🔄 Converting Gemini data to DataFrame...")
        
        # Extract headers and data
        headers = gemini_data.get('table_info', {}).get('headers', [])
        rows_data = gemini_data.get('data', [])
        
        if not headers or not rows_data:
            raise ValueError("No headers or data found in Gemini response")
        
        # Create DataFrame
        df_data = []
        for row in rows_data:
            cells = row.get('cells', [])
            # Ensure we have the right number of columns
            if len(cells) == len(headers):
                df_data.append(cells)
            else:
                # Pad with empty strings if columns don't match
                padded_cells = cells + [''] * (len(headers) - len(cells))
                df_data.append(padded_cells[:len(headers)])
        
        # Create DataFrame
        df = pd.DataFrame(df_data, columns=headers)
        
        logger.info(f"   ✅ DataFrame created: {df.shape[0]} rows × {df.shape[1]} columns")
        return df
        
    except Exception as e:
        logger.error(f"❌ Error creating DataFrame: {e}")
        raise

def enhance_dataframe(df: pd.DataFrame):
    """
    Enhance DataFrame with additional formatting and data cleaning.
    """
    try:
        logger.info("   🔧 Enhancing DataFrame...")
        
        # Clean column names
        df.columns = [col.strip() for col in df.columns]
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        # Clean cell data
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.strip()
                # Replace empty strings with NaN
                df[col] = df[col].replace('', pd.NA)
        
        # Reset index
        df = df.reset_index(drop=True)
        
        logger.info(f"   ✅ DataFrame enhanced: {df.shape[0]} rows × {df.shape[1]} columns")
        return df
        
    except Exception as e:
        logger.error(f"❌ Error enhancing DataFrame: {e}")
        raise

def save_to_excel(df: pd.DataFrame, output_path: str):
    """
    Save DataFrame to Excel with professional formatting.
    """
    try:
        logger.info(f"   💾 Saving to Excel: {os.path.basename(output_path)}")
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write DataFrame
            df.to_excel(writer, sheet_name='Table Data', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Table Data']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max width 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format header row
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add borders to all cells
            from openpyxl.styles import Border, Side
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = thin_border
        
        logger.info(f"   ✅ Excel file saved successfully")
        
    except Exception as e:
        logger.error(f"❌ Error saving Excel file: {e}")
        raise

def create_summary_report(df: pd.DataFrame, gemini_data: dict, output_dir: str):
    """
    Create a comprehensive summary report.
    """
    try:
        logger.info("   📝 Creating summary report...")
        
        # Generate summary statistics
        total_rows = len(df)
        total_columns = len(df.columns)
        
        # Count non-empty cells
        non_empty_cells = df.count().sum()
        total_cells = total_rows * total_columns
        data_completeness = (non_empty_cells / total_cells) * 100 if total_cells > 0 else 0
        
        # Create summary
        summary_path = os.path.join(output_dir, "excel_creation_summary.txt")
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write("Excel Table Creation Summary\n")
            f.write("=" * 35 + "\n\n")
            f.write(f"Source: Gemini API Extraction\n")
            f.write(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(f"Table Dimensions:\n")
            f.write(f"  - Total Rows: {total_rows}\n")
            f.write(f"  - Total Columns: {total_columns}\n")
            f.write(f"  - Total Cells: {total_cells}\n\n")
            f.write(f"Data Quality:\n")
            f.write(f"  - Non-empty Cells: {non_empty_cells}\n")
            f.write(f"  - Data Completeness: {data_completeness:.1f}%\n\n")
            f.write(f"Column Headers:\n")
            for i, col in enumerate(df.columns, 1):
                f.write(f"  {i}. {col}\n")
            f.write(f"\nSample Data (First 3 rows):\n")
            f.write("-" * 30 + "\n")
            f.write(df.head(3).to_string(index=False))
            f.write(f"\n\nOutput Files:\n")
            f.write(f"  - Excel File: Table Data\n")
            f.write(f"  - Summary: {os.path.basename(summary_path)}\n")
        
        logger.info(f"   ✅ Summary report saved: {summary_path}")
        
    except Exception as e:
        logger.error(f"❌ Error creating summary report: {e}")
        raise

def main():
    """Main function for creating Excel from Gemini data."""
    parser = argparse.ArgumentParser(description='Create Excel table from Gemini API data')
    parser.add_argument('gemini_data', help='Path to Gemini API JSON data file')
    parser.add_argument('--output', '-o', default='data/output/excel_from_gemini',
                       help='Output directory for Excel file and reports')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.gemini_data):
        logger.error(f"❌ Gemini data file not found: {args.gemini_data}")
        sys.exit(1)
    
    try:
        # Create output directory
        os.makedirs(args.output, exist_ok=True)
        
        logger.info(f"🚀 Starting Excel Creation from Gemini API Data...")
        logger.info(f"   Gemini Data: {os.path.basename(args.gemini_data)}")
        logger.info(f"   Output Directory: {args.output}")
        
        # Step 1: Load Gemini data
        gemini_data = load_gemini_data(args.gemini_data)
        
        # Step 2: Convert to DataFrame
        df = create_dataframe_from_gemini(gemini_data)
        
        # Step 3: Enhance DataFrame
        df_enhanced = enhance_dataframe(df)
        
        # Step 4: Save to Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"gemini_extracted_table_{timestamp}.xlsx"
        excel_path = os.path.join(args.output, excel_filename)
        
        save_to_excel(df_enhanced, excel_path)
        
        # Step 5: Create summary report
        create_summary_report(df_enhanced, gemini_data, args.output)
        
        # Display results
        logger.info(f"\n🎉 Excel Creation Complete!")
        logger.info(f"   Excel File: {excel_filename}")
        logger.info(f"   Total Rows: {len(df_enhanced)}")
        logger.info(f"   Total Columns: {len(df_enhanced.columns)}")
        logger.info(f"   Output Directory: {args.output}")
        
        # Show sample data
        logger.info(f"\n📊 Sample Data Preview:")
        logger.info(f"   Headers: {', '.join(df_enhanced.columns[:5])}{'...' if len(df_enhanced.columns) > 5 else ''}")
        logger.info(f"   First Row: {list(df_enhanced.iloc[0][:5])}{'...' if len(df_enhanced.columns) > 5 else ''}")
        
    except Exception as e:
        logger.error(f"❌ Excel creation failed: {e}")
        raise

if __name__ == "__main__":
    main()
